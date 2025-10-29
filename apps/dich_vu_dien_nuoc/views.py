from django.shortcuts import render, get_object_or_404
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_protect

from django.forms.models import model_to_dict
from django.http import JsonResponse, HttpResponse
from django.utils import timezone

from django.db import transaction
from django.db.models import Sum, Q, F, Case, When, FloatField, Min, Max
from django.db.models.functions import Coalesce
from django.contrib.auth.decorators import login_required

from .models import *

from json import loads, dumps
from io import BytesIO
from django.template.loader import render_to_string
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from datetime import datetime,timedelta
import json
from num2words import num2words

import calendar
from urllib.parse import quote
# Create your views here.

# ------------------------------------ VIEW BÁO CÁO DOANH THU ------------------------------------

def get_filtered_revenue_data(start_date=None, end_date=None, customer_id=None, service_id=None):
    """Hàm logic trung tâm để lấy tất cả dữ liệu báo cáo từ database"""

    # 1. Xây dựng bộ lọc cơ sở
    filters = Q()
    if start_date and end_date:
        filters = Q(id_thanhtoan_dichvu__thoigiantao__date__range=[start_date, end_date])
    if customer_id and customer_id != 'all':
        filters &= Q(id_thanhtoan_dichvu__id_hopdong_id=customer_id)
    if service_id and service_id != 'all':
        filters &= Q(id_dichvu_id=service_id)

    # Lấy queryset cơ sở đã lọc
    ct_thanhtoan_qs = CtThanhtoanDichvu.objects.filter(filters).select_related(
        'id_thanhtoan_dichvu__id_hopdong', 'id_dichvu'
    )
    
    # 2. Lấy dữ liệu cho Thẻ tóm tắt (Summary Cards)
    # Lấy ID các thông báo thanh toán duy nhất từ kết quả đã lọc
    thanhtoan_ids = ct_thanhtoan_qs.values_list('id_thanhtoan_dichvu_id', flat=True).distinct()
    thanhtoan_qs = ThanhtoanDichvu.objects.filter(id__in=thanhtoan_ids)
    
    summary_agg_thanhtoan_qs = thanhtoan_qs.aggregate(
        total_revenue=Coalesce(Sum('tongtiensauthue') - Sum(F('giamtru')/100 * F('tongtientruocthue')), 0.0)
    )

    summary_agg_ct_thanhtoan_qs = ct_thanhtoan_qs.aggregate(
        total_revenue_dien=Coalesce(Sum('tiensauthue', filter=Q(id_dichvu__id_loaidichvu_id=19)), 0.0),
        total_revenue_nuoc=Coalesce(Sum('tiensauthue', filter=Q(id_dichvu__id_loaidichvu_id=20)), 0.0),
    )

    summary = {
        "total_revenue": summary_agg_thanhtoan_qs['total_revenue'],
        "total_revenue_dien": summary_agg_ct_thanhtoan_qs['total_revenue_dien'],
        "total_revenue_nuoc": summary_agg_ct_thanhtoan_qs['total_revenue_nuoc']
    }

    # 3. Lấy dữ liệu Bảng thông báo
    thong_bao = list(thanhtoan_qs.order_by('-thoigiantao').values(
        'sotbdv', 'tongtiensauthue', 'thoigiantao',
        tencongty=F('id_hopdong__tencongty'),
        giam_tru = F('giamtru')/100 * F('tongtientruocthue'),
        tongtienthanhtoan = F('tongtiensauthue') - F('giamtru')/100 * F('tongtientruocthue')
    ).order_by("tencongty", "-sotbdv"))

    # 4. Lấy dữ liệu Bảng chi tiết dịch vụ
    chi_tiet_dich_vu_raw = list(ct_thanhtoan_qs
        .values('id_thanhtoan_dichvu__id_hopdong__tencongty', 'id_dichvu__tendichvu', 'id_dichvu_id')
        .annotate(
            donvitinh=Min('donvitinh'),
            tongtiensauthue=Sum('tiensauthue'),
            tongsosudung=Sum('sosudung')
        ).order_by('id_thanhtoan_dichvu__id_hopdong__tencongty')
    )
    
    # Gom nhóm lại cho frontend
    chi_tiet_dich_vu = {}
    for item in chi_tiet_dich_vu_raw:
        ten_cong_ty = item['id_thanhtoan_dichvu__id_hopdong__tencongty']
        if ten_cong_ty not in chi_tiet_dich_vu:
            chi_tiet_dich_vu[ten_cong_ty] = {
                'tencongty': ten_cong_ty,
                'tong_tien_dich_vu': 0,
                'dich_vu': {}
            }
        chi_tiet_dich_vu[ten_cong_ty]['dich_vu'][item['id_dichvu_id']] = item
        chi_tiet_dich_vu[ten_cong_ty]['tong_tien_dich_vu'] += item['tongtiensauthue']
        
    # 5. Lấy dữ liệu Bảng tổng hợp Điện - Nước (ID loại dịch vụ 19: Điện, 20: Nước)
    dien_nuoc_qs = ct_thanhtoan_qs.filter(id_dichvu__id_loaidichvu_id__in=[19, 20])
    tong_hop_dien_nuoc = list(dien_nuoc_qs
        .values(tencongty=F('id_thanhtoan_dichvu__id_hopdong__tencongty'))
        .annotate(
            tong_tien_dien=Sum(Case(When(id_dichvu__id_loaidichvu_id=19, then=F('tiensauthue')), default=0.0, output_field=FloatField())),
            so_dien=Sum(Case(When(id_dichvu__id_loaidichvu_id=19, then=F('sosudung')), default=0.0, output_field=FloatField())),
            tong_tien_nuoc=Sum(Case(When(id_dichvu__id_loaidichvu_id=20, then=F('tiensauthue')), default=0.0, output_field=FloatField())),
            so_nuoc=Sum(Case(When(id_dichvu__id_loaidichvu_id=20, then=F('sosudung')), default=0.0, output_field=FloatField())),
        ).order_by('tencongty')
    )

    return {
        "summary": summary,
        "data": {
            "thong_bao": thong_bao,
            "chi_tiet_dich_vu": list(chi_tiet_dich_vu.values()),
            "tong_hop_dien_nuoc": tong_hop_dien_nuoc,
        }
    }

@login_required
def view_bao_cao_doanh_thu(request):
    """Hiển thị báo cáo doanh thu lần đầu."""
    
    # Lấy dữ liệu ban đầu
    initial_data = get_filtered_revenue_data()

    # Chuẩn bị dữ liệu danh sách dịch vụ và tính tổng
    danh_sach_dich_vu = {}
    for congty in initial_data['data']['chi_tiet_dich_vu']:
        for value_dv in congty.get("dich_vu", {}).values():

            # Nếu dịch vụ chưa có trong danh sách, khởi tạo nó
            if value_dv['id_dichvu_id'] not in danh_sach_dich_vu:
                danh_sach_dich_vu[value_dv['id_dichvu_id']] = {
                    "id_dichvu": value_dv['id_dichvu_id'],
                    "tendichvu": value_dv['id_dichvu__tendichvu'],
                    "total_amount": 0, 
                    "total_usage": 0,
                    "unit": value_dv['donvitinh']
                }
            
            # Cập nhật tổng tiền và tổng số lượng sử dụng
            danh_sach_dich_vu[value_dv['id_dichvu_id']]['total_amount'] += value_dv['tongtiensauthue']
            danh_sach_dich_vu[value_dv['id_dichvu_id']]['total_usage'] += value_dv['tongsosudung']

    context = {
        "danh_sach_khach_thue": Hopdong.objects.all().order_by("tencongty"),
        "danh_sach_dich_vu": sorted(danh_sach_dich_vu.values(), key=lambda x: x['tendichvu']),
        
        "tong_doanh_thu": initial_data['summary']['total_revenue'],
        "tong_doanh_thu_dien": initial_data['summary'].get('total_revenue_dien', None),
        "tong_doanh_thu_nuoc": initial_data['summary'].get('total_revenue_nuoc', None),
        
        "thong_bao_data": initial_data['data']['thong_bao'],
        "tong_hop_chi_tiet_dich_vu": initial_data['data']['chi_tiet_dich_vu'],
        "tong_hop_chi_tiet_dich_vu_dien_nuoc": initial_data['data']['tong_hop_dien_nuoc'],

        # TÍNH TOÁN CÁC GIÁ TRỊ TỔNG CỘNG
        'total_tiengomthue': sum(item.get('tongtiensauthue', 0) for item in initial_data['data']['thong_bao']),
        'total_giamtru': sum(item.get('giam_tru', 0) for item in initial_data['data']['thong_bao']),
        'total_tongtientt': sum(item.get('tongtienthanhtoan', 0) for item in initial_data['data']['thong_bao']),

        'tota_tiendichvu': sum(item.get('tong_tien_dich_vu', 0) for item in initial_data['data']['chi_tiet_dich_vu']),

        'total_tongtiendien': sum(float(c.get('tong_tien_dien', 0)) for c in initial_data['data']['tong_hop_dien_nuoc']),
        'total_sodiensudung': sum(float(c.get('so_dien', 0)) for c in initial_data['data']['tong_hop_dien_nuoc']),
        'total_tongtiennuoc': sum(float(c.get('tong_tien_nuoc', 0)) for c in initial_data['data']['tong_hop_dien_nuoc']),
        'total_sonuocsudung': sum(float(c.get('so_nuoc', 0)) for c in initial_data['data']['tong_hop_dien_nuoc']),
    }

    return render(request, "dich_vu_dien_nuoc/bao_cao_doanh_thu.html", context)

@login_required
def api_bao_cao_doanh_thu_filter(request):
    """API để lọc báo cáo doanh thu."""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    customer_id = request.GET.get('customer')
    service_id = request.GET.get('service')
    
    data = get_filtered_revenue_data(start_date, end_date, customer_id, service_id)
    data['success'] = True # Thêm trường success cho JS
    
    return JsonResponse(data)

@login_required
def api_bao_cao_doanh_thu_export(request):
    """
    API để xuất báo cáo doanh thu ra file Excel với 3 sheet.
    """
    # 1. Lấy dữ liệu đã được tối ưu
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    customer_id = request.GET.get('customer')
    service_id = request.GET.get('service')
    
    # Tái sử dụng hàm logic chung để lấy toàn bộ dữ liệu cần thiết
    report_data = get_filtered_revenue_data(start_date, end_date, customer_id, service_id)
    data = report_data['data']

    # 2. Tạo Workbook và các Sheet trong bộ nhớ
    workbook = openpyxl.Workbook()
    
    # --- Sheet 1: Danh sách Thông báo ---
    sheet1 = workbook.active
    sheet1.title = "Danh sách Thông báo"
    headers1 = ["Mã TB", "Khách thuê", "Tổng tiền gồm thuế", "Giảm trừ", "Tổng tiền TT", "Ngày tạo"]
    sheet1.append(headers1)
    
    for item in data['thong_bao']:
        row = [
            item['sotbdv'],
            item['tencongty'],
            item['tongtiensauthue'],
            item['giam_tru'],
            item['tongtienthanhtoan'],
            item['thoigiantao'].strftime('%d/%m/%Y %H:%M') if item.get('thoigiantao') else ''
        ]
        sheet1.append(row)
        
    # --- Sheet 2: Chi tiết Dịch vụ ---
    sheet2 = workbook.create_sheet(title="Chi tiết Dịch vụ")
    
    # Lấy danh sách dịch vụ để làm header động
    all_services = {
        s['id_dichvu_id']: s['id_dichvu__tendichvu'] 
        for company in data['chi_tiet_dich_vu'] 
        for s in company['dich_vu'].values()
    }
    sorted_service_ids = sorted(all_services.keys())
    
    headers2 = ["Công ty"] + [all_services[sid] for sid in sorted_service_ids] + ['Tổng cộng']
    sheet2.append(headers2)

    # In đậm header "Tổng cộng"
    sheet2.cell(row=1, column=len(headers2)).font = Font(bold=True)

    for company in data['chi_tiet_dich_vu']:
        row = [company['tencongty']]
        for service_id in sorted_service_ids:
            service_data = company['dich_vu'].get(service_id)
            # Ghi cả thành tiền và số lượng sử dụng
            cell_value = service_data['tongtiensauthue'] if service_data else ""
            row.append(cell_value)
        row.append(company['tong_tien_dich_vu'])
        sheet2.append(row)

        # In đậm ô Tổng cộng
        sheet2.cell(row=sheet2.max_row, column=len(row)).font = Font(bold=True)

    # --- Sheet 3: Tổng hợp Điện - Nước ---
    sheet3 = workbook.create_sheet(title="Tổng hợp Điện - Nước")
    headers3 = ["Công ty", "Tổng tiền điện", "Số điện sử dụng (kWh)", "Tổng tiền nước", "Số nước sử dụng (m³)"]
    sheet3.append(headers3)
    
    for item in data['tong_hop_dien_nuoc']:
        row = [
            item['tencongty'],
            item['tong_tien_dien'],
            item['so_dien'],
            item['tong_tien_nuoc'],
            item['so_nuoc'],
        ]
        sheet3.append(row)

    # 3. Tối ưu độ rộng cột cho tất cả sheet
    for sheet in workbook.worksheets:
        for col in sheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    # 4. Lưu file vào buffer và trả về response
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"BaoCaoDoanhThu_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ------------------------------------ VIEW QUẢN LÝ DỊCH VỤ ------------------------------------
@login_required
@ensure_csrf_cookie
def view_danh_sach_thong_bao (request):
    return render(request, "dich_vu_dien_nuoc/danh_sach_thong_bao.html")

@login_required
def api_danh_sach_thong_bao(request):
    """API để lấy danh sách thông báo dịch vụ"""
    try:
        # Lấy tham số filter từ request
        month = request.GET.get('month', '')
        year = request.GET.get('year', '')
        company = request.GET.get('company', '')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        
        # Query cơ bản - join với bảng HopDong để lấy tên công ty và chú thích
        queryset = ThanhtoanDichvu.objects.select_related('id_hopdong').all()
        
        # Áp dụng filters
        if month:
            queryset = queryset.filter(thoigiantao__month=int(month))
            
        if year:
            queryset = queryset.filter(thoigiantao__year=int(year))
            
        if company:
            # Tìm theo tên công ty trong bảng HopDong
            queryset = queryset.filter(
                id_hopdong__tencongty__icontains=company
            )
        
        # Sắp xếp theo thời gian tạo từ mới đến cũ
        queryset = queryset.order_by('-thoigiantao')
        
        # Tính tổng số bản ghi
        total_count = queryset.count()
        
        # Phân trang
        start_index = (page - 1) * per_page
        end_index = start_index + per_page
        paginated_queryset = queryset[start_index:end_index]
        
        # Chuẩn bị data trả về
        notifications = []
        for index, item in enumerate(paginated_queryset, start=start_index + 1):
            # Lấy thông tin từ hợp đồng liên kết
            hop_dong = item.id_hopdong
            
            notifications.append({
                'id': item.id,
                'stt': index,
                # Số TBDV từ bảng ThanhToan_DichVu
                'sotbdv': item.sotbdv if item.sotbdv else f'TBDV-{item.id:06d}',
                
                # Tên công ty từ bảng HopDong thông qua id_hopdong
                'tencongty': hop_dong.tencongty if hop_dong and hop_dong.tencongty else 'Chưa có tên công ty',
                
                # Chú thích từ bảng HopDong thông qua id_hopdong
                'chuthich': hop_dong.chuthich if hop_dong and hop_dong.chuthich else '',
                
                # Ngày tạo từ ThoigianTao trong bảng ThanhToan_DichVu
                'ngaytao': item.thoigiantao.strftime('%Y-%m-%d') if item.thoigiantao else '',
                'ngaytao_display': item.thoigiantao.strftime('%d/%m/%Y') if item.thoigiantao else '',
                
                # Tổng tiền trước thuế và sau thuế từ bảng ThanhToan_DichVu
                'tongtientruocthue': float(item.tongtientruocthue) if item.tongtientruocthue else 0,
                'tongtiensauthue': float(item.tongtiensauthue) if item.tongtiensauthue else 0,
                
                # Thông tin bổ sung
                'thoigiantao': item.thoigiantao.isoformat() if item.thoigiantao else None,
                'id_hopdong': hop_dong.id_hopdong if hop_dong else None,
                'sohd': hop_dong.sohd if hop_dong and hop_dong.sohd else '',
            })
        
        return JsonResponse({
            'success': True,
            'data': notifications,
            'pagination': {
                'current_page': page,
                'per_page': per_page,
                'total_count': total_count,
                'total_pages': max(1, (total_count + per_page - 1) // per_page)
            }
        })
        
    except Exception as e:
        import traceback
        print(f"API Error: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e),
            'data': [],
            'pagination': {'current_page': 1, 'per_page': 10, 'total_count': 0, 'total_pages': 1}
        }, status=500)

@login_required
def api_danh_sach_cong_ty(request):
    """API để lấy danh sách công ty cho dropdown filter"""
    try:
        companies = Hopdong.objects.values_list('tencongty', flat=True).distinct().order_by('tencongty')
        companies_list = [company for company in companies if company]  # Loại bỏ giá trị None
        
        return JsonResponse({
            'success': True,
            'companies': companies_list
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@csrf_protect
@require_http_methods(["DELETE"])
def api_xoa_thong_bao(request, notification_id):
    """API để xóa thông báo"""
    if request.method == 'DELETE':
        try:
            notification = ThanhtoanDichvu.objects.get(id=notification_id)
            notification.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Xóa thông báo thành công'
            })
            
        except ThanhtoanDichvu.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Không tìm thấy thông báo'
            }, status=404)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'error': 'Method not allowed'
    }, status=405)

@login_required
def api_danh_sach_cong_ty_chua_tao(request):
    try:
        # SỬA: Bỏ tham số day, chỉ lọc theo month và year
        month = int(request.GET.get('month', 0))
        year = int(request.GET.get('year', 0))
        
        if not month or not year:
            return JsonResponse({
                'success': False,
                'error': 'Thiếu tham số month hoặc year'
            }, status=400)
        
        # Lấy tất cả công ty
        all_companies = list(Hopdong.objects.values_list('tencongty', flat=True).distinct().order_by('tencongty'))
        all_companies = [company for company in all_companies if company and company.strip()]
        
        # SỬA: Tìm các công ty đã tạo thông báo cho THÁNG cụ thể (bỏ filter theo day)
        created_companies = list(ThanhtoanDichvu.objects.filter(
            thoigiantao__year=year,
            thoigiantao__month=month,
            id_hopdong__tencongty__isnull=False
        ).values_list('id_hopdong__tencongty', flat=True).distinct())
        
        created_companies = [company for company in created_companies if company and company.strip()]
        
        # Lọc ra các công ty CHƯA tạo thông báo cho tháng này
        available_companies = [company for company in all_companies 
                             if company not in created_companies]
        
        return JsonResponse({
            'success': True,
            'companies': available_companies,
            'debug': {
                'period': f'{month}/{year}',
                'all_companies_count': len(all_companies),
                'created_companies': created_companies,
                'available_count': len(available_companies)
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def api_lay_dich_vu_ky_truoc(request):
    """API để lấy dịch vụ từ kỳ thanh toán trước của công ty"""
    try:
        company = request.GET.get('company', '')
        year = int(request.GET.get('year', 0))
        month = int(request.GET.get('month', 0))
        
        if not company or not year or not month:
            return JsonResponse({
                'success': False,
                'error': 'Thiếu tham số bắt buộc'
            }, status=400)
        
        # Tính kỳ thanh toán trước
        if month == 1:
            prev_year = year - 1
            prev_month = 12
        else:
            prev_year = year
            prev_month = month - 1
        
        # Tìm thông báo của kỳ trước
        prev_notification = ThanhtoanDichvu.objects.filter(
            id_hopdong__tencongty=company,
            thoigiantao__year=prev_year,
            thoigiantao__month=prev_month
        ).first()
        
        services = []
        previous_discount = 0

        if prev_notification:
            # THÊM: Lấy phần trăm giảm trừ từ kỳ trước
            previous_discount = float(prev_notification.giamtru) if prev_notification.giamtru else 0
            
            # Lấy chi tiết dịch vụ từ kỳ trước
            prev_services = CtThanhtoanDichvu.objects.filter(
                id_thanhtoan_dichvu=prev_notification
            ).select_related('id_dichvu')
            
            for service in prev_services:
                # Chỉ số mới của kỳ trước = chỉ số cũ của kỳ hiện tại
                old_reading = service.chisomoi if service.chisomoi is not None else None
                
                services.append({
                    'service_id': service.id_dichvu.id_dichvu if service.id_dichvu else None,
                    'name': service.tendichvu or (service.id_dichvu.tendichvu if service.id_dichvu else ''),
                    'unit': service.donvitinh or '',
                    'old_reading': old_reading,
                    'new_reading': '',  # Để trống cho user nhập
                    'factor': service.heso or 1,
                    'unit_price': service.dongia or 0,
                    'usage': service.sosudung or 0 if old_reading is None else '',  # Nếu không có chỉ số thì copy số sử dụng
                    'tax_rate': service.loaithue or 8
                })
        
        return JsonResponse({
            'success': True,
            'services': services,
            'previous_discount': previous_discount  # THÊM: Trả về phần trăm giảm trừ từ kỳ trước
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def api_danh_sach_tat_ca_dich_vu(request):
    """API để lấy tất cả dịch vụ cho dropdown"""
    try:
        services = Dichvu.objects.all().order_by('tendichvu')
        
        services_list = []
        for service in services:
            services_list.append({
                'id': service.id_dichvu,
                'name': service.tendichvu or f'Dịch vụ {service.id_dichvu}'
            })
        
        return JsonResponse({
            'success': True,
            'services': services_list
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@csrf_protect
@require_http_methods(["POST"])
def api_tao_moi_thong_bao(request):
    """API để tạo mới thông báo dịch vụ"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Method not allowed'
        }, status=405)
    
    try:
        data = json.loads(request.body)
        
        company = data.get('company', '').strip()
        period = data.get('period', {})
        sotbdv = data.get('sotbdv', '').strip()
        discount = float(data.get('discount', 0))
        services_data = data.get('services', [])
        
        # Validation
        if not company:
            return JsonResponse({
                'success': False,
                'error': 'Tên công ty không được để trống'
            }, status=400)
        
        # SỬA: Kiểm tra period format mới
        if not period or not period.get('full_date'):
            return JsonResponse({
                'success': False,
                'error': 'Kỳ thanh toán không hợp lệ'
            }, status=400)
        
        if not sotbdv:
            return JsonResponse({
                'success': False,
                'error': 'Số TBDV không được để trống'
            }, status=400)
        
        if not services_data:
            return JsonResponse({
                'success': False,
                'error': 'Phải có ít nhất một dịch vụ'
            }, status=400)
        
        # SỬA: Parse full_date thay vì year/month riêng lẻ
        try:
            thoigian_tao = datetime.strptime(period['full_date'], '%Y-%m-%d')
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Định dạng ngày không hợp lệ'
            }, status=400)
        
        # SỬA: Kiểm tra trùng lặp theo ngày/tháng/năm
        existing = ThanhtoanDichvu.objects.filter(
            id_hopdong__tencongty=company,
            thoigiantao__year=thoigian_tao.year,
            thoigiantao__month=thoigian_tao.month
        ).exists()
        
        if existing:
            return JsonResponse({
                'success': False,
                'error': f'Công ty {company} đã có thông báo cho ngày {thoigian_tao.strftime("%d/%m/%Y")}'
            }, status=400)
        
        # Lấy hợp đồng của công ty
        hop_dong = Hopdong.objects.filter(tencongty=company).first()
        if not hop_dong:
            return JsonResponse({
                'success': False,
                'error': f'Không tìm thấy hợp đồng cho công ty {company}'
            }, status=400)
        
        # Kiểm tra số TBDV đã tồn tại chưa
        existing_sotbdv = ThanhtoanDichvu.objects.filter(sotbdv=sotbdv).exists()
        if existing_sotbdv:
            return JsonResponse({
                'success': False,
                'error': f'Số TBDV {sotbdv} đã tồn tại'
            }, status=400)
        
        # Tính tổng tiền
        total_before_tax = 0
        total_after_tax = 0
        
        for service_data in services_data:
            usage = float(service_data.get('usage', 0))
            unit_price = float(service_data.get('unit_price', 0))
            tax_rate = float(service_data.get('tax_rate', 8))
            
            before_tax = usage * unit_price
            after_tax = before_tax * (1 + tax_rate / 100)
            
            total_before_tax += before_tax
            total_after_tax += after_tax
        
        # Áp dụng giảm trừ
        discount_amount = total_before_tax * (discount / 100) if discount > 0 else 0
        final_amount = total_after_tax - discount_amount
        
        # Tạo thông báo chính
        thong_bao = ThanhtoanDichvu.objects.create(
            thoigiantao=thoigian_tao,  # SỬA: Dùng ngày được chọn thay vì datetime.now()
            sotbdv=sotbdv,
            id_hopdong=hop_dong,
            giamtru=discount,
            tongtientruocthue=total_before_tax,
            tongtiensauthue=total_after_tax,
        )
        
        # Tạo chi tiết dịch vụ (code không thay đổi)
        for service_data in services_data:
            service_id = service_data.get('service_id')
            if not service_id:
                continue
            
            try:
                dich_vu = Dichvu.objects.get(id_dichvu=service_id)
            except Dichvu.DoesNotExist:
                continue
            
            usage = float(service_data.get('usage', 0))
            unit_price = float(service_data.get('unit_price', 0))
            tax_rate = float(service_data.get('tax_rate', 8))
            
            tien_truoc_thue = usage * unit_price
            thue = tien_truoc_thue * tax_rate / 100
            tien_sau_thue = tien_truoc_thue + thue
            
            CtThanhtoanDichvu.objects.create(
                id_dichvu=dich_vu,
                tientruocthue=tien_truoc_thue,
                thue=thue,
                tiensauthue=tien_sau_thue,
                donvitinh=service_data.get('unit', ''),
                chisocu=service_data.get('old_reading') if service_data.get('old_reading') else None,
                chisomoi=service_data.get('new_reading') if service_data.get('new_reading') else None,
                heso=int(service_data.get('factor', 1)),
                dongia=unit_price,
                sosudung=usage,
                loaithue=tax_rate,
                id_thanhtoan_dichvu=thong_bao,
                tendichvu=dich_vu.tendichvu
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Tạo thông báo thành công',
            'notification_id': thong_bao.id
        })
        
    except Exception as e:
        import traceback
        print(f"API Error: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
@login_required
def api_chi_tiet_thong_bao(request, notification_id):
    """API để lấy chi tiết thông báo cho chỉnh sửa"""
    try:
        thong_bao = get_object_or_404(ThanhtoanDichvu.objects.select_related('id_hopdong'), id=notification_id)
        
        # Lấy chi tiết dịch vụ
        chi_tiet_services = CtThanhtoanDichvu.objects.filter(
            id_thanhtoan_dichvu=thong_bao
        ).select_related('id_dichvu').order_by('id')
        
        services_list = []
        for ct in chi_tiet_services:
            services_list.append({
                'original_id': ct.id,  # ID gốc để update
                'service_id': ct.id_dichvu.id_dichvu if ct.id_dichvu else None,
                'name': ct.tendichvu or (ct.id_dichvu.tendichvu if ct.id_dichvu else ''),
                'unit': ct.donvitinh or '',
                'old_reading': ct.chisocu if ct.chisocu is not None else '',
                'new_reading': ct.chisomoi if ct.chisomoi is not None else '',
                'factor': ct.heso or 1,
                'unit_price': float(ct.dongia) if ct.dongia else 0,
                'usage': float(ct.sosudung) if ct.sosudung else 0,
                'tax_rate': float(ct.loaithue) if ct.loaithue else 8
            })
        
        # Format period
        period_str = f"{thong_bao.thoigiantao.year}-{thong_bao.thoigiantao.month:02d}" if thong_bao.thoigiantao else ""
        
        notification_data = {
            'id': thong_bao.id,
            'company': thong_bao.id_hopdong.tencongty if thong_bao.id_hopdong else '',
            'period': period_str,
            'sotbdv': thong_bao.sotbdv or '',
            'discount': float(thong_bao.giamtru) if thong_bao.giamtru else 0,
            'services': services_list,
            'total_before_tax': float(thong_bao.tongtientruocthue) if thong_bao.tongtientruocthue else 0,
            'total_after_tax': float(thong_bao.tongtiensauthue) if thong_bao.tongtiensauthue else 0
        }
        
        return JsonResponse({
            'success': True,
            'notification': notification_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@csrf_protect
@require_http_methods(["PUT"])
def api_cap_nhat_thong_bao(request, notification_id):
    """API để cập nhật thông báo"""
    if request.method != 'PUT':
        return JsonResponse({
            'success': False,
            'error': 'Method not allowed'
        }, status=405)
    
    try:
        data = json.loads(request.body)
        
        thong_bao = get_object_or_404(ThanhtoanDichvu, id=notification_id)
        
        discount = float(data.get('discount', 0))
        services_data = data.get('services', [])
        
        if not services_data:
            return JsonResponse({
                'success': False,
                'error': 'Phải có ít nhất một dịch vụ'
            }, status=400)
        
        # Tính tổng tiền
        total_before_tax = 0
        total_after_tax = 0
        
        for service_data in services_data:
            usage = float(service_data.get('usage', 0))
            unit_price = float(service_data.get('unit_price', 0))
            tax_rate = float(service_data.get('tax_rate', 8))
            
            before_tax = usage * unit_price
            after_tax = before_tax * (1 + tax_rate / 100)
            
            total_before_tax += before_tax
            total_after_tax += after_tax
        
        # Áp dụng giảm trừ
        discount_amount = total_before_tax * (discount / 100) if discount > 0 else 0
        final_amount = total_after_tax - discount_amount
        
        # Cập nhật thông báo chính
        thong_bao.giamtru = discount
        thong_bao.tongtientruocthue = total_before_tax
        thong_bao.tongtiensauthue = total_after_tax
        thong_bao.save()
        
        # Xóa chi tiết dịch vụ cũ
        CtThanhtoanDichvu.objects.filter(id_thanhtoan_dichvu=thong_bao).delete()
        
        # Tạo lại chi tiết dịch vụ
        for service_data in services_data:
            service_id = service_data.get('service_id')
            if not service_id:
                continue
            
            try:
                dich_vu = Dichvu.objects.get(id_dichvu=service_id)
            except Dichvu.DoesNotExist:
                continue
            
            usage = float(service_data.get('usage', 0))
            unit_price = float(service_data.get('unit_price', 0))
            tax_rate = float(service_data.get('tax_rate', 8))
            
            tien_truoc_thue = usage * unit_price
            thue = tien_truoc_thue * tax_rate / 100
            tien_sau_thue = tien_truoc_thue + thue
            
            CtThanhtoanDichvu.objects.create(
                id_dichvu=dich_vu,
                tientruocthue=tien_truoc_thue,
                thue=thue,
                tiensauthue=tien_sau_thue,
                donvitinh=service_data.get('unit', ''),
                chisocu=service_data.get('old_reading') if service_data.get('old_reading') else None,
                chisomoi=service_data.get('new_reading') if service_data.get('new_reading') else None,
                heso=int(service_data.get('factor', 1)),
                dongia=unit_price,
                sosudung=usage,
                loaithue=tax_rate,
                id_thanhtoan_dichvu=thong_bao,
                tendichvu=dich_vu.tendichvu
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Cập nhật thông báo thành công'
        })
        
    except Exception as e:
        import traceback
        print(f"API Error: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def api_in_thong_bao(request, notification_id):
    """API để tạo HTML in cho thông báo"""
    try:
        thong_bao = get_object_or_404(ThanhtoanDichvu.objects.select_related('id_hopdong'), id=notification_id)
        
        # Lấy chi tiết dịch vụ
        chi_tiet_services = CtThanhtoanDichvu.objects.filter(
            id_thanhtoan_dichvu=thong_bao
        ).select_related('id_dichvu').order_by('id')
        
        # Tính toán thời gian
        created_date = thong_bao.thoigiantao
        day = created_date.day
        month = created_date.month
        year = created_date.year
        
        # Tính ngày cuối tháng của kỳ thanh toán
        last_day_of_month = calendar.monthrange(year, month)[1]
        payment_day = last_day_of_month
        payment_month = month
        payment_year = year
        
        # Lấy tên công ty và tạo tên file với tháng/năm
        company_name = thong_bao.id_hopdong.tencongty if thong_bao.id_hopdong else 'Khong_co_ten'
        
        # Loại bỏ ký tự đặc biệt và tạo tên file an toàn
        safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_company_name = safe_company_name.replace(' ', '_')
        
        # Tạo tên file đề xuất với format: TBDV - Tên công ty - MM/YYYY
        suggested_filename = f"TBDV - {company_name} - T{month:02d}/{year}"
        
        # Tạo các dòng dịch vụ
        services_rows = ""
        total_before_tax_calc = 0
        total_tax_calc = 0
        total_after_tax_calc = 0
        
        for index, service in enumerate(chi_tiet_services, 1):
            old_reading = service.chisocu if service.chisocu is not None else ''
            new_reading = service.chisomoi if service.chisomoi is not None else ''
            
            before_tax = float(service.tientruocthue or 0)
            tax = float(service.thue or 0)
            after_tax = float(service.tiensauthue or 0)
            
            total_before_tax_calc += before_tax
            total_tax_calc += tax
            total_after_tax_calc += after_tax
            
            service_name = service.tendichvu
            if not service_name and service.id_dichvu:
                service_name = service.id_dichvu.tendichvu
            if not service_name:
                service_name = f'Dịch vụ {service.id}'
            
            services_rows += f'''
                    <tr>
                        <td>{index}</td>
                        <td class="left">{service_name}</td>
                        <td>{service.donvitinh or ''}</td>
                        <td class="right">{format_number_vn(old_reading) if old_reading != '' else ''}</td>
                        <td class="right">{format_number_vn(new_reading) if new_reading != '' else ''}</td>
                        <td class="right">{format_number_vn(service.heso) if service.heso else "1.00"}</td>
                        <td class="right">{format_number_vn(service.dongia) if service.dongia else "0.00"}</td>
                        <td class="right">{format_number_vn(service.sosudung) if service.sosudung else "0.00"}</td>
                        <td class="right">{format_number_vn(before_tax)}</td>
                        <td class="right">{format_number_vn(tax)}</td>
                        <td class="right">{format_number_vn(after_tax)}</td>
                    </tr>'''
        
        # Tính toán giảm trừ và tiền cuối cùng
        discount_amount = 0
        if thong_bao.giamtru and thong_bao.giamtru > 0:
            discount_amount = total_before_tax_calc * (thong_bao.giamtru / 100)
        
        final_amount = total_after_tax_calc - discount_amount
        
        # Chuyển đổi tiền thành chữ
        amount_in_words = num2words(int(final_amount), lang='vi').capitalize() + " đồng"
        
        # Đọc template HTML
        
        
        context = {
            'day': day,
            'month': month,
            'year': year,
            'period_month': created_date.month,
            'period_year': created_date.year,
            'payment_day': payment_day,
            'payment_month': payment_month,
            'payment_year': payment_year,
            'company_name': company_name,
            'suggested_filename': suggested_filename,  # Tên file có tháng/năm
            'services_rows': services_rows,
            'total_before_tax': format_number_vn(total_before_tax_calc),
            'tax_amount': format_number_vn(total_tax_calc),
            'total_after_tax': format_number_vn(total_after_tax_calc),
            'discount_amount': format_number_vn(discount_amount) if discount_amount > 0 else "0.00",
            'final_amount': format_number_vn(final_amount),
            'amount_in_words': amount_in_words
        }
        
        html_content = render_to_string('dich_vu_dien_nuoc/notification_print_template.html', context)
        
        response = HttpResponse(html_content, content_type='text/html; charset=utf-8')
        
        # Thêm header để browser hiểu filename (cần encode để tránh lỗi với ký tự đặc biệt)
        from urllib.parse import quote
        encoded_filename = quote(suggested_filename.encode('utf-8'))
        response['Content-Disposition'] = f'inline; filename="{encoded_filename}.pdf"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def api_in_nhieu_thong_bao(request):
    """API để tạo nhiều file in"""
    try:
        ids_str = request.GET.get('ids', '')
        if not ids_str:
            return JsonResponse({
                'success': False,
                'error': 'Không có ID thông báo'
            }, status=400)
        
        ids = [int(id.strip()) for id in ids_str.split(',') if id.strip().isdigit()]
        if not ids:
            return JsonResponse({
                'success': False,
                'error': 'ID không hợp lệ'
            }, status=400)
        
        # Tạo URL cho từng thông báo
        print_urls = []
        for notification_id in ids:
            url = f'/dichvudiennuoc/api/in-thong-bao/{notification_id}/'
            print_urls.append(url)
        
        return JsonResponse({
            'success': True,
            'print_urls': print_urls,
            'count': len(print_urls)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def format_number_vn(number):
    """Format số theo chuẩn: phẩy cho hàng nghìn, chấm cho thập phân"""
    if number is None or number == 0:
        return "0.00"
    
    # Làm tròn đến 1 chữ số thập phân
    number = round(float(number), 2)
    
    # Tách phần nguyên và phần thập phân
    integer_part = int(number)
    decimal_part = number - integer_part
    
    # Format phần nguyên với dấu phẩy cho hàng nghìn (giữ nguyên dấu phẩy mặc định)
    formatted_integer = f"{integer_part:,}"
    
    # Nếu có phần thập phân và != 0
    if decimal_part > 0:
        decimal_str = f"{decimal_part:.2f}"[2:]  # Lấy phần sau dấu chấm
        return f"{formatted_integer}.{decimal_str}"  # Dùng chấm cho thập phân
    else:
        return formatted_integer
    
def view_quan_ly_loai_dich_vu (request):
    """Hiển thị danh sách loại dịch vụ"""

    loai_dich_vu_list = Loaidichvu.objects.all().order_by("id_loaidichvu")

    dich_vu_list = Dichvu.objects.all().order_by("id_dichvu").select_related("id_loaidichvu")

    context = {
        "danh_sach_dich_vu": dich_vu_list,
        "danh_sach_loai_dich_vu": loai_dich_vu_list
    }

    return render(request, "dich_vu_dien_nuoc/quan_ly_loai_dich_vu.html", context)

@login_required
@csrf_protect
@require_POST
@transaction.atomic
def api_dich_vu_update_or_create(request):
    """Cập nhật hoặc tạo mới dịch vụ"""    
    try:
        data = loads(request.body)
        pk = data.get('id_dichvu', None)

        # Biến lưu lại dữ liệu được cập nhật hoặc tạo mới để response lại cho giao diện
        dich_vu_instance = None
    
        if pk:
            # Cập nhật dịch vụ
            dich_vu_instance = get_object_or_404(Dichvu, id_dichvu=pk)
            
            dich_vu_instance.id_loaidichvu_id = data.get("id_loaidichvu", None)
            dich_vu_instance.tendichvu = data.get("tendichvu", None)
            dich_vu_instance.chuthich = data.get("chuthich", None)
            dich_vu_instance.ngayghi = timezone.now()
            
            # Lưu lại các thay đổi
            dich_vu_instance.save()
        
        else:
            # Tạo mới dịch vụ
            data.pop("id_dichvu", None)

            dich_vu_instance = Dichvu.objects.create(
                id_loaidichvu_id=data.get("id_loaidichvu", None),
                tendichvu=data.get("tendichvu", None),
                chuthich=data.get("chuthich", None),
                ngayghi=timezone.now()
            )
        
        return JsonResponse({'success': True, 'message': 'Lưu loại dịch vụ thành công!', 'data': model_to_dict(dich_vu_instance)})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': f"Dữ liệu thêm mới chưa đúng - {str(e)}"}, status=400)
    

@login_required
@csrf_protect
@require_POST
@transaction.atomic
def api_loai_dich_vu_update_or_create(request):
    """Cập nhật hoặc tạo mới Loại dịch vụ"""    
    try:
        data = loads(request.body)
        pk = data.get('id_loaidichvu', None)

        # Biến lưu lại dữ liệu được cập nhật hoặc tạo mới để response lại cho giao diện
        loai_dich_vu_instance = None

        if pk:
            # Cập nhật dịch vụ
            loai_dich_vu_instance = get_object_or_404(Loaidichvu, id_loaidichvu=pk)
            
            data.pop("id_dichvu", None) # Xóa id_dichvu khỏi data để tránh lỗi khi cập nhật
            for key, value in data.items():
                setattr(loai_dich_vu_instance, key, value)
            
            # Lưu lại các thay đổi
            loai_dich_vu_instance.save()
        
        else:
            # Tạo mới dịch vụ
            data.pop("id_loaidichvu", None)

            loai_dich_vu_instance = Loaidichvu.objects.create(
                tenloaidichvu=data.get("tenloaidichvu", None),
                chuthich=data.get("chuthich", None),
            )
        
        return JsonResponse({'success': True, 'message': 'Lưu loại dịch vụ thành công!', 'data': model_to_dict(loai_dich_vu_instance)})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': f"Dữ liệu thêm mới chưa đúng - {str(e)}"}, status=400)


@login_required
@csrf_protect
def api_dich_vu_delete(request, pk):
    """Xóa dịch vụ"""        
    try: 
        dichvu = get_object_or_404(Dichvu, id_dichvu = pk)
        dichvu.delete()
        return JsonResponse({'success': True, 'message': 'Xóa dịch vụ thành công!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
@csrf_protect
def api_loai_dich_vu_delete(request, pk):
    """Xóa loại dịch vụ"""
    try: 
        loaidichvu = get_object_or_404(Loaidichvu, id_loaidichvu = pk)
        loaidichvu.delete()
        return JsonResponse({'success': True, 'message': 'Xóa dịch vụ thành công!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


# ------------------------------------ VIEW QUẢN LÝ KHÁCH THUÊ ------------------------------------

@login_required
def view_quan_ly_khach_thue (request):
    """Hiển thị danh sách hợp đồng thuê và các dịch vụ liên quan"""
    hopdong_list = Hopdong.objects.prefetch_related("hopdongdichvu_set__id_dichvu").order_by("-ngaytaohopdong")

    response_data = []
    for hopdong in hopdong_list:
        dichvu_list = [
            {
                "id_hopdongdichvu": dv.id_hopdongdichvu,
                "id_dichvu": dv.id_dichvu_id,
                "tendichvu": dv.id_dichvu.tendichvu,
                "donvitinh": dv.donvitinh,
                "dongia": dv.dongia,
                "chuthich": dv.chuthich
            }
            for dv in hopdong.hopdongdichvu_set.all()
        ]

        response_data.append({
            "id_hopdong": hopdong.id_hopdong,
            "tencongty": hopdong.tencongty,
            "kythanhtoan": hopdong.kythanhtoan,
            "sohd": hopdong.sohd,
            "chuthich": hopdong.chuthich,
            "ngaytaohopdong": hopdong.ngaytaohopdong,
            "tiencoc": hopdong.tiencoc,
            "ngayketthuc": hopdong.ngayketthuc if hopdong.ngayketthuc else "",
            "ngaybatdautinhtien": hopdong.ngaybatdautinhtien if hopdong.ngaybatdautinhtien else "",
            "trangthai": hopdong.trangthai,
            "dichvu_list": dumps(dichvu_list)
        })
    
    context = {
        "danh_sach_hop_dong": response_data
    }

    return render(request, "dich_vu_dien_nuoc/quan_ly_khach_thue.html", context)

@login_required
@csrf_protect
@transaction.atomic
@require_POST
def api_quan_ly_khach_thue_update_or_create(request):
    try:
        data = loads(request.body)
        dichvu_list = data.pop('dichvu_list', [])
        hopdong_id = data.get('id_hopdong')

        # Chuyển đổi các trường ngày tháng và rỗng
        for field in ['ngaytaohopdong', 'ngayketthuc', 'ngaybatdautinhtien']:
            if data.get(field) == '':
                data[field] = None

        # Format dữ liệu số
        data['tiencoc'] = float(data.get('tiencoc', 0)) if data.get('tiencoc') else None
        
        # Logic cập nhật hoặc tạo mới Hợp đồng
        if hopdong_id:
            hopdong_instance = get_object_or_404(Hopdong, id_hopdong=hopdong_id)
            # Bỏ id_hopdong khỏi data trước khi cập nhật
            data.pop('id_hopdong', None)

            # Cập nhật các trường hợp có trong payload
            for key, value in data.items():
                setattr(hopdong_instance, key, value)
            hopdong_instance.save()
        else:
            data.pop('id_hopdong', None)
            hopdong_instance = Hopdong.objects.create(**data)

        # Lấy ra những dịch vụ gửi lên từ payload
        submitted_dichvu_ids = {item.get('id_hopdongdichvu') for item in dichvu_list if item.get('id_hopdongdichvu')}
        
        # Xóa những dịch vụ không có trong payload gửi lên
        HopdongDichvu.objects.filter(id_hopdong=hopdong_instance).exclude(id_hopdongdichvu__in=submitted_dichvu_ids).delete()

        # Cập nhật hoặc tạo mới dịch vụ
        for service_data in dichvu_list:
            if not service_data.get('id_dichvu'):
                continue

            hopdong_dichvu_id = service_data.get('id_hopdongdichvu')
            
            defaults = {
                'id_dichvu_id': service_data['id_dichvu'],
                'donvitinh': service_data.get('donvitinh'),
                'dongia': service_data.get('dongia', 0) if service_data.get("dongia") else None,
                'chuthich': service_data.get('chuthich')
            }

            if hopdong_dichvu_id:
                # Cập nhật
                HopdongDichvu.objects.filter(id_hopdongdichvu=hopdong_dichvu_id).update(**defaults)
            else:
                # Tạo mới
                HopdongDichvu.objects.create(
                    id_hopdong=hopdong_instance,
                    **defaults
                )
        
        return JsonResponse({'success': True, 'message': 'Lưu hợp đồng thành công!'})

    except Exception as e:
        return JsonResponse({'success': False, 'message': f"Dữ liệu thêm mới chưa đúng - {str(e)}"}, status=400)

@login_required
@csrf_protect
@transaction.atomic
@require_POST
def api_quan_ly_khach_thue_delete(request, pk):
    try:
        hopdong = get_object_or_404(Hopdong, id_hopdong=pk)
        hopdong.delete()
        return JsonResponse({'success': True, 'message': 'Xóa hợp đồng thành công!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
def api_get_all_services (request):
    dich_vu = Dichvu.objects.all().values("id_dichvu", "tendichvu")
    return JsonResponse(list(dich_vu), safe=False)